import openpyxl
import shutil

# For scanID
source = "/storage/research/cinn_comp/ThalSR/zan/derivatives/test/practice/scanID"
# To sessionID
dest = "/storage/research/cinn_comp/ThalSR/zan/derivatives/test/practice/sessionID"
# Location of excel file
excFile = "/storage/research/cinn_comp/ThalSR/zan/FTHP_Metadata.xlsx"

wb = openpyxl.load_workbook(excFile)
ws = wb.get_sheet_by_name('Sheet1')

#UNFINISHED!! strugglint to make it work
# iterate through excel and display data 
for row in ws.iter_rows(min_row=2, min_col=1, max_col=2): 
    for cell in row: 
        print(cell.value, end=" ") 
    print()

# need to iterate through the selected rows then connect them to the
# existing folders in source and destination, then move source to destination folders
# based on the excel cols
    
for row in ws.iter_rows():
    if row.value == source.value:
        file = source + row.value
        fileDest = dest + row.value
        shutil.copyfile(file, fileDest)
